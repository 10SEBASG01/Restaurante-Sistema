import json
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, Count
from django.db.models.functions import TruncDate, ExtractIsoWeekDay, ExtractHour
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse

# Librerías profesionales para la exportación de documentos
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# Guardián de seguridad y modelos relacionales del sistema
from apps.usuarios.views import requiere_permiso
from apps.auditoria.models import Auditoria

# 🔥 MODELOS DE TU MÓDULO DE FACTURACIÓN IMPORTADOS
from apps.facturacion.models import Factura, FacturaDetalle

# =====================================================================
# 📊 VISTA PRINCIPAL: DASHBOARD DE REPORTES
# =====================================================================
@login_required
@requiere_permiso('modulo_reportes')
def dashboard_reportes(request):
    """
    🎯 DASHBOARD DE INTELIGENCIA DE NEGOCIO (BI)
    
    Calcula y renderiza las métricas clave de rendimiento (KPIs) del restaurante 
    basándose en el historial inmutable de las Facturas emitidas. Genera datos para
    gráficos de ventas, rentabilidad por categorías y estadísticas de concurrencia.
    """

    # ==========================================
    # 📅 1. CONTROL DE RANGO DE FECHAS
    # ==========================================
    # Por defecto, el dashboard muestra la actividad de los últimos 7 días.
    hoy = timezone.now().date()
    hace_una_semana = hoy - timedelta(days=6)
    
    fecha_inicio_str = request.GET.get('inicio')
    fecha_fin_str = request.GET.get('fin')
    
    desde = hace_una_semana
    hasta = hoy
    
    # Intenta convertir las fechas del formulario HTML a objetos Date de Python
    if fecha_inicio_str and fecha_fin_str:
        try:
            desde = timezone.datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            hasta = timezone.datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        except ValueError:
            pass # Ignora entradas de fecha corruptas

    # ==========================================
    # 🔍 2. EXTRACCIÓN DE DATOS (QUERIES)
    # ==========================================
    # Filtra directamente sobre las FACTURAS EMITIDAS en el rango de fechas establecido
    facturas_filtradas = Factura.objects.filter(
        fecha_emision__date__range=(desde, hasta)
    )
    
    total_pedidos = facturas_filtradas.count()
    
    # Cruce relacional: Obtiene los platillos exactos vendidos dentro de esas facturas
    detalles_reporte = FacturaDetalle.objects.filter(id_factura__in=facturas_filtradas)
    
    # ==========================================
    # 💰 3. CÁLCULO DE KPIs GLOBALES
    # ==========================================
    # Sumatorias automáticas interactuando con los montos inmutables de la factura
    conceptos_globales = facturas_filtradas.aggregate(
        ventas_totales=Sum('total')
    )
    unidades_globales = detalles_reporte.aggregate(
        productos_vendidos=Sum('cantidad')
    )
    
    ventas_totales = conceptos_globales['ventas_totales'] or 0
    productos_vendidos = unidades_globales['productos_vendidos'] or 0
    # Evita el error de división por cero si no hay ventas
    ticket_promedio = ventas_totales / total_pedidos if total_pedidos > 0 else 0

    # ==========================================
    # 📈 4. DATOS PARA GRÁFICOS (VENTAS DIARIAS)
    # ==========================================
    # Agrupa las facturas por día exacto y suma el total recaudado en cada uno
    ventas_diarias_query = facturas_filtradas.annotate(
        dia=TruncDate('fecha_emision')
    ).values('dia').annotate(
        total=Sum('total')
    ).order_by('dia')
    
    chart_labels = []
    chart_data = []
    for registro in ventas_diarias_query:
        chart_labels.append(registro['dia'].strftime('%d/%m'))
        chart_data.append(float(registro['total']))

    # ==========================================
    # 🍕 5. RENTABILIDAD POR CATEGORÍA
    # ==========================================
    # Multiplica cantidad * precio_historico y agrupa por el nombre de la categoría del platillo
    ventas_categorias_query = detalles_reporte.values(
        'id_platillo__id_categoria__nombre_categoria'
    ).annotate(
        total_categoria=Sum(F('cantidad') * F('precio_unitario_historico'))
    ).order_by('-total_categoria')
    
    categorias_lista = []
    for item in ventas_categorias_query:
        total_cat = item['total_categoria'] or 0
        porcentaje = (total_cat / ventas_totales * 100) if ventas_totales > 0 else 0
        
        categorias_lista.append({
            'nombre': item['id_platillo__id_categoria__nombre_categoria'],
            'total': total_cat,
            'porcentaje': round(porcentaje, 1)
        })

    # ==========================================
    # 🏆 6. ESTADÍSTICAS RÁPIDAS (TOP RENDIMIENTO)
    # ==========================================
    mejor_dia = "Sin datos"
    mejor_hora = "Sin datos"
    producto_estrella = "Sin datos"

    if total_pedidos > 0:
        # Producto Estrella (El más pedido según el historial facturado)
        producto_top = detalles_reporte.values('id_platillo__nombre_platillo').annotate(
            total_vendido=Sum('cantidad')
        ).order_by('-total_vendido').first()
        if producto_top:
            producto_estrella = producto_top['id_platillo__nombre_platillo']

        # Horario con mayor concurrencia (Agrupación por hora del día)
        hora_top = facturas_filtradas.annotate(
            hora=ExtractHour('fecha_emision')
        ).values('hora').annotate(
            cantidad_pedidos=Count('id_factura')
        ).order_by('-cantidad_pedidos').first()
        if hora_top and hora_top['hora'] is not None:
            h = hora_top['hora']
            mejor_hora = f"{h:02d}:00 - {(h+1)%24:02d}:00"

        # Día de la semana con mayor recaudación económica (1=Lunes, 7=Domingo)
        dia_top = facturas_filtradas.annotate(
            dia_semana=ExtractIsoWeekDay('fecha_emision')
        ).values('dia_semana').annotate(
            plata_dia=Sum('total')
        ).order_by('-plata_dia').first()
        
        dias_espanol = {1: 'Lunes', 2: 'Martes', 3: 'Miércoles', 4: 'Jueves', 5: 'Viernes', 6: 'Sábado', 7: 'Domingo'}
        if dia_top and dia_top['dia_semana']:
            mejor_dia = dias_espanol.get(dia_top['dia_semana'], "Sin datos")

    # ==========================================
    # 📦 7. EMPAQUETADO Y RENDERIZADO
    # ==========================================
    # Traemos los últimos 5 registros de auditoría para el panel lateral
    actividad_reciente = Auditoria.objects.all()[:5]

    context = {
        'ventas_totales': ventas_totales,
        'total_pedidos': total_pedidos,
        'ticket_promedio': ticket_promedio,
        'productos_vendidos': productos_vendidos,
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
        'categorias_lista': categorias_lista,
        'desde': desde.strftime('%d/%m/%Y'),
        'hasta': hasta.strftime('%d/%m/%Y'),
        'desde_input': desde.strftime('%Y-%m-%d'),
        'hasta_input': hasta.strftime('%Y-%m-%d'),
        'mejor_dia': mejor_dia,
        'mejor_hora': mejor_hora,
        'producto_estrella': producto_estrella,
        'actividad_reciente': actividad_reciente,
    }
    return render(request, 'reportes/dashboard.html', context)


# =====================================================================
# 📊 EXPORTAR A EXCEL (OPENPYXL)
# =====================================================================
@login_required
@requiere_permiso('modulo_reportes')
def exportar_excel(request):
    """
    🎯 EXPORTACIÓN DE REPORTES A EXCEL
    
    Genera un archivo '.xlsx' descargable con el detalle tabular de las facturas 
    emitidas en un rango de fechas. Aplica estilos corporativos y auto-ajusta 
    las columnas para una presentación profesional.
    """
    # 1. Filtro de fechas (Misma lógica del dashboard)
    fecha_inicio_str = request.GET.get('inicio')
    fecha_fin_str = request.GET.get('fin')
    
    hoy = timezone.now().date()
    desde = hoy - timedelta(days=6)
    hasta = hoy
    
    if fecha_inicio_str and fecha_fin_str:
        try:
            desde = timezone.datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            hasta = timezone.datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Consultamos las facturas validadas
    facturas = Factura.objects.filter(
        fecha_emision__date__range=(desde, hasta)
    ).order_by('-fecha_emision')

    # 2. Inicialización del Libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    # Estilos de la cabecera (Fondo dorado y letra blanca)
    header_font = Font(bold=True, color="FFFFFF", name="Segoe UI")
    header_fill = PatternFill(start_color="C49A45", end_color="C49A45", fill_type="solid")
    
    headers = ['N° Factura', 'Fecha y Hora', 'Cajero Responsable', 'Cliente', 'Total Cobrado']
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 25

    # 3. Llenado iterativo de datos
    row_num = 2
    total_general = 0
    
    for fac in facturas:
        total_general += fac.total
        
        ws.cell(row=row_num, column=1, value=fac.secuencial).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=2, value=fac.fecha_emision.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row_num, column=3, value=fac.id_cajero.get_full_name() or fac.id_cajero.username)
        ws.cell(row=row_num, column=4, value=fac.cliente_nombre)
        
        # Formato moneda para los montos
        cell_monto = ws.cell(row=row_num, column=5, value=float(fac.total))
        cell_monto.number_format = '"$"#,##0.00'
        cell_monto.alignment = Alignment(horizontal='right')
        
        row_num += 1

    # 4. Fila de Totalización
    ws.cell(row=row_num, column=4, value="TOTAL GENERAL FACTURADO:").font = Font(bold=True, name="Segoe UI")
    ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='right')
    
    cell_total = ws.cell(row=row_num, column=5, value=float(total_general))
    cell_total.font = Font(bold=True, color="C49A45", name="Segoe UI")
    cell_total.number_format = '"$"#,##0.00'
    cell_total.alignment = Alignment(horizontal='right')

    # 5. Auto-ajuste inteligente del ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 4

    # 6. Preparación de la respuesta HTTP para descarga
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Ventas_{desde}_al_{hasta}.xlsx"'
    wb.save(response)
    return response


# =====================================================================
# 📄 EXPORTAR A PDF (REPORTLAB)
# =====================================================================
@login_required
@requiere_permiso('modulo_reportes')
def exportar_pdf(request):
    """
    🎯 EXPORTACIÓN DE REPORTES A PDF
    
    Genera un documento PDF estructurado y listo para impresión, detallando las 
    ventas facturadas en un período específico usando la librería ReportLab.
    """
    # 1. Filtro de fechas (Misma lógica del dashboard)
    fecha_inicio_str = request.GET.get('inicio')
    fecha_fin_str = request.GET.get('fin')
    
    hoy = timezone.now().date()
    desde = hoy - timedelta(days=6)
    hasta = hoy
    
    if fecha_inicio_str and fecha_fin_str:
        try:
            desde = timezone.datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            hasta = timezone.datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    facturas = Factura.objects.filter(
        fecha_emision__date__range=(desde, hasta)
    ).order_by('-fecha_emision')

    # 2. Preparación del Documento PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Reporte_Ventas_{desde}_al_{hasta}.pdf"'

    doc = SimpleDocTemplate(
        response, 
        pagesize=letter, 
        rightMargin=40, 
        leftMargin=40, 
        topMargin=40, 
        bottomMargin=40,
        title="Reporte de Ventas",
        author="Sabor & Arte"
    )
    elementos = []
    estilos = getSampleStyleSheet()

    # Estructura del Título y Subtítulo del documento
    estilo_titulo = estilos['Title']
    estilo_titulo.fontName = 'Helvetica-Bold'
    estilo_titulo.fontSize = 22
    estilo_titulo.textColor = colors.HexColor('#1e293b')
    estilo_titulo.alignment = 0 

    titulo = Paragraph("Sabor & Arte - Reporte de Ventas", estilo_titulo)
    subtitulo = Paragraph(f"Período de análisis: {desde.strftime('%d/%m/%Y')} al {hasta.strftime('%d/%m/%Y')} | Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}", estilos['Normal'])
    elementos.append(titulo)
    elementos.append(subtitulo)
    elementos.append(Spacer(1, 25))

    # 3. Construcción de la Tabla de Datos
    datos_tabla = [['N° Factura', 'Fecha y Hora', 'Cajero Responsable', 'Total']]
    total_general = 0

    for fac in facturas:
        total_general += fac.total
        
        datos_tabla.append([
            fac.secuencial,
            fac.fecha_emision.strftime('%d/%m/%Y %H:%M'),
            fac.id_cajero.get_full_name() or fac.id_cajero.username,
            f"${fac.total:.2f}"
        ])

    # Fila de Totalización
    datos_tabla.append(['', '', 'TOTAL GENERAL:', f"${total_general:.2f}"])

    # 4. Asignación de Estilos a la Matriz de la Tabla
    tabla = Table(datos_tabla, colWidths=[100, 120, 205, 90])
    
    estilo_tabla = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c49a45')), 
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'), 
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#ffffff')),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#e2e8f0')),
        ('FONTNAME', (2, -1), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (3, -1), (3, -1), colors.HexColor('#c49a45')),
        ('TOPPADDING', (0, -1), (-1, -1), 12),
    ])
    tabla.setStyle(estilo_tabla)
    elementos.append(tabla)

    # 5. Generación Final del Documento
    doc.build(elementos)
    return response